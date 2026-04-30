import pandas as pd
from pandas import DataFrame, MultiIndex, Series, Index

from config import PatientDir, PATHS
from model_comparison.correlation_of_prediction_errors import correlation_of_prediction_errors_from_pdir

PER_PTNT_METRIC_NAMES = ['total_clips', 'preictal_clips', 'non_preictal_clips']
SPLITS = ['train', 'test']


def _save_styled_excel(styler, df: DataFrame, path,
                       autosize_index: bool = True, autosize_data: bool = True) -> None:
    """Save styled DataFrame to Excel, optionally auto-sizing index and/or data columns."""
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        styler.to_excel(writer)
        ws = writer.sheets['Sheet1']

        if autosize_index:
            for level in range(df.index.nlevels):
                index_vals = df.index.get_level_values(level)
                index_name = '' if df.index.names[level] is None else str(df.index.names[level])
                max_len = max([len(index_name), *(len(str(v)) for v in index_vals)])
                ws.column_dimensions[get_column_letter(level + 1)].width = max_len

        if autosize_data:
            data_col_offset = df.index.nlevels
            for i, col in enumerate(df.columns):
                if isinstance(col, tuple):
                    width = max(len(str(part)) for part in col if part is not None)
                else:
                    width = len(str(col))
                ws.column_dimensions[get_column_letter(data_col_offset + i + 1)].width = width


def _load_ptnt_metrics(pdir: PatientDir, split: str, model: str) -> tuple[Series, Series]:
    """
    Load and partition the saved metrics for a given patient, split, and model.
    :return: (per_patient metrics, per_model metrics)
    """
    m: Series = pd.read_pickle(pdir.model_eval_subdir(split, model).metrics_table.pickle)
    m = m.drop(['model', 'data_split'])
    # Split into per patient and per model metrics
    per_ptnt = m[PER_PTNT_METRIC_NAMES].astype(int)
    per_model = m.drop(PER_PTNT_METRIC_NAMES).astype(float)
    return per_ptnt, per_model


def make_comparison_table(
        pdirs: list[PatientDir],
        models: tuple[str, ...] = ('CNN', 'ensemble'),
) -> tuple[DataFrame, DataFrame]:
    """
    Create a comparison table per patient and per model that compare various metrics in between patients, models, and
    data splits.
    :param pdirs: List of patient directories to include.
    :param models: Model names to compare.
    :return: per patient table, per model table
    """
    if not pdirs:
        raise ValueError("pdirs must not be empty.")

    ptnts = [pdir.name for pdir in pdirs]

    # Cache all metric loads to avoid redundant disk I/O
    cache: dict[tuple, tuple[Series, Series]] = {}
    for pdir in pdirs:
        for split in SPLITS:
            for model in models:
                cache[(pdir.name, split, model)] = _load_ptnt_metrics(pdir, split, model)

    # Use the first entry to discover the column structure
    per_ptnt_probe, per_model_probe = cache[(pdirs[0].name, SPLITS[0], models[0])]

    # Build per-patient comparison table
    per_p = DataFrame(
        index=Index(ptnts, name='patient'),
        columns=MultiIndex.from_product([per_ptnt_probe.index, SPLITS], names=['metric', 'split']),
        dtype='Int64'
    )

    # Fill per-patient table
    for pdir in pdirs:
        for split in SPLITS:
            per_ptnt, _ = cache[(pdir.name, split, models[0])]
            for metric in per_ptnt.index:
                per_p.loc[pdir.name, (metric, split)] = per_ptnt[metric]

    # Add preictal_ratio and corr_of_pred_errors for each split with separate loops for insertion order
    for split in SPLITS:
        per_p[('preictal_ratio', split)] = per_p[('preictal_clips', split)] / per_p[('total_clips', split)]
    for split in SPLITS:
        # Compute correlation of prediction errors for each patient/split
        per_p[('corr_of_pred_errors', split)] = [correlation_of_prediction_errors_from_pdir(pdir, split) for pdir in
                                                 pdirs]

    # Build per-model comparison table
    per_m = DataFrame(
        index=MultiIndex.from_product([ptnts, models], names=['patient', 'model']),
        columns=MultiIndex.from_product([per_model_probe.index, SPLITS], names=['metric', 'split']),
        dtype=float
    )

    for pdir in pdirs:
        for split in SPLITS:
            for model in models:
                _, per_model = cache[(pdir.name, split, model)]
                per_m.loc[(pdir.name, model), (per_model.index, split)] = per_model.values

    return per_p, per_m


def make_comparison_table_and_save(
        pdirs: list[PatientDir],
        models: tuple[str, ...] = ('CNN', 'ensemble'),
) -> None:
    """Build comparison tables and persist them in multiple formats."""
    per_ptnt, per_model = make_comparison_table(pdirs, models)

    # Style per_model
    metric_level = per_model.columns.get_level_values('metric')

    neutral_metrics = {'best_threshold'}
    lower_better_metrics = {'rel_tifw', 'p_hanley_mcneil'}
    higher_better_metrics = set(metric_level) - neutral_metrics - lower_better_metrics

    neutral_cols = per_model.columns[metric_level.isin(neutral_metrics)]
    lower_better_cols = per_model.columns[metric_level.isin(lower_better_metrics)]
    higher_better_cols = per_model.columns[metric_level.isin(higher_better_metrics)]

    per_model_styled = per_model.style
    if len(higher_better_cols) > 0:
        per_model_styled = per_model_styled.background_gradient(cmap='RdYlGn', vmin=0, vmax=1, axis=None,
                                                                subset=higher_better_cols)
    if len(neutral_cols) > 0:
        per_model_styled = per_model_styled.background_gradient(cmap='Blues', vmin=0, vmax=1, axis=None,
                                                                subset=neutral_cols)
    if len(lower_better_cols) > 0:
        per_model_styled = per_model_styled.background_gradient(cmap='RdYlGn_r', vmin=0, vmax=1, axis=None,
                                                                subset=lower_better_cols)

    # Style per_ptnt
    preictal_cols = [col for col in per_ptnt.columns if col[0] == 'preictal_ratio']
    cope_cols = [col for col in per_ptnt.columns if col[0] == 'corr_of_pred_errors']
    per_ptnt_styled = (per_ptnt.style
                       .background_gradient(cmap='Blues', vmin=0, vmax=per_ptnt[preictal_cols].max().max(), axis=None,
                                            subset=preictal_cols)
                       .background_gradient(cmap='RdYlGn', vmin=0, vmax=1, axis=None, subset=cope_cols)
                       )

    # Save DataFrames
    PATHS.statistical_results_dir.mkdir(parents=True, exist_ok=True)

    per_model.to_pickle(PATHS.per_model_comparison_table.pickle)
    _save_styled_excel(per_model_styled, per_model, PATHS.per_model_comparison_table.xlsx, autosize_data=False)

    per_ptnt.to_pickle(PATHS.per_patient_comparison_table.pickle)
    _save_styled_excel(per_ptnt_styled, per_ptnt, PATHS.per_patient_comparison_table.xlsx, autosize_data=False,)


if __name__ == '__main__':
    pdirs_ = PATHS.patient_dirs(include_fake_ptnts=False)
    make_comparison_table_and_save(pdirs_)
    print(f"Created comparison tables for {len(pdirs_)} patients.")
