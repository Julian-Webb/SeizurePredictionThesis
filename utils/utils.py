import logging
import os
import time
from functools import wraps
from pathlib import Path

import numpy as np
from pandas import Series, DataFrame


def safe_float_to_int(num: float) -> int:
    if num != int(num):
        raise ValueError(f"Number {num} has decimal values")
    return int(num)


class FunctionTimer:
    def __init__(self, label: str = "unknown_func"):
        self.label = label
        self.start = None
        self.elapsed = None

    def __enter__(self):
        self.start = time.perf_counter()
        return self

    def __exit__(self, exc_type, exc, tb):
        self.elapsed = time.perf_counter() - self.start
        logging.info(f"[TIMING] {self.label}: {self.elapsed:.3f}s")


def timeit(
        f=None,
        *,
        arg_indices: list[int] | None = None,
        kwarg_names: list[str] | None = None,
        show_all: bool = False,
):
    """Decorator that logs the execution time of a function.

    Usage:
        @timeit                                             – shows no args/kwargs
        @timeit(show_all=True)                             – shows all args & kwargs
        @timeit(arg_indices=[0, 1])                         – selected positional args by index
        @timeit(kwarg_names=['patient'])                    – selected keyword args by name
        @timeit(arg_indices=[0], kwarg_names=['patient'])   – both
    """

    def decorator(func):
        @wraps(func)
        def wrap(*args, **kw):
            start = time.perf_counter()
            result = func(*args, **kw)
            elapsed = time.perf_counter() - start

            selected = []
            if show_all:
                selected.extend(repr(a) for a in args)
                selected.extend(f"{k}={repr(v)}" for k, v in kw.items())
            else:
                if arg_indices is not None:
                    selected.extend(repr(args[i]) for i in arg_indices if i < len(args))
                if kwarg_names is not None:
                    selected.extend(f"{k}={repr(kw[k])}" for k in kwarg_names if k in kw)

            args_str = f"{', '.join(selected)}" if selected else ""
            logging.info(f"[TIMING] {elapsed:.3f}s - {func.__name__}({args_str})")
            return result

        return wrap

    # Supports both @timeit and @timeit(...) usage
    if f is not None:
        return decorator(f)
    return decorator


def import_tensorflow_with_available_gpus(available_gpus: list[int]):
    gpus_str = ','.join(map(str, available_gpus))
    import os
    os.environ['CUDA_VISIBLE_DEVICES'] = gpus_str
    # noinspection PyUnusedImports
    import tensorflow as tf


MAC_PATTERNS = [
    '._.DS_Store',  # Resource fork for .DS_Store
    '.DS_Store',  # Finder metadata
    '._*',  # All resource fork files
    '.AppleDouble',  # Apple Double format directory
    '.LSOverride'  # Finder custom attributes
]


def clean_mac_files(directory: Path):
    """Removes macOS system files, such as .DS_Store and ._ files"""
    # go through all the files in the directory and delete them
    # logging.info('===== Removing macOS system files =====')
    removed_files = []
    for root, dirs, files in os.walk(directory):
        for filename in files:
            # Check exact matches
            path = Path(root, filename)
            if filename in MAC_PATTERNS:
                try:
                    path.unlink()
                    removed_files.append(path)
                    # logging.info(f"Removed: {Path(root, filename)}")
                except OSError as e:
                    logging.warning(f"Error removing {path}: {e}")

            # Check pattern matches (for ._*)
            elif filename.startswith('._'):
                try:
                    path.unlink()
                    removed_files.append(path)
                    # logging.info(f"Removed: {path}")
                except OSError as e:
                    logging.warning(f"Error removing {path}: {e}")

    logging.info(f"Removed {len(removed_files)} mac system files")


def contains_nan(obj: DataFrame | Series | np.ndarray):
    if isinstance(obj, DataFrame) or isinstance(obj, Series):
        obj = obj.values
    return np.isnan(obj).any()


def autofit_excel_columns(
        xlsx_path: str,
        sheet_name: str | None = None,
        max_width: int = 24,
        min_width: int = 8,
        scale: float = 0.75,
) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        width = max(min_width, min(int(max_len * scale), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(xlsx_path)

